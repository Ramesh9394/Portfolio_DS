import openpyxl
from openpyxl import workbook
from openpyxl import load_workbook

from datetime import date


# startmodel = load_workbook('C4-A20-15D_xxxxxx_AFOxxxx_A00_20220531.xlsx')
#
#
# st_number = input('Please enter the Bemi Nr: ')
# afo = input("please enter the AFO number: ")
# st_name = input("Please Enter name of the station: ")
# project = input("please Enter name of the project: ")
# mirror = input("Please type 'Yes' if station is mirror otherwise 'No': ")
#
# today = date.today()
#
# # Accessing the all sheets and writing the required information
# for i,sheets in enumerate(startmodel.worksheets):
#     print(sheets)
#     sheets['W65'] = i
#     sheets['W67'] = today
#     sheets['W64'] = project
#     sheets["N69"] = project
#     sheets["N65"] = afo
#     if mirror == 'Yes':
#         sheets["Y67"] = st_number + '\n' + st_number
#     else:
#         sheets["Y67"] = st_number

# Loading the data file
template = load_workbook('template.xlsx')
print(template.sheetnames)

# Let's identify the maximum rows present in the master data sheet for all BG
ws = template.active
for i in range(3,200):
    cellvalue = ws.cell(row=i, column=2).value
    if cellvalue == None:
        max_row = i
        break
print("Max_rows are: ", max_row)

rows = ws.iter_rows(min_row=3, max_row=max_row - 1, min_col=1, max_col=6)
print(rows)

# Saving all the values in data structures for using later
bg_set = set()
bemi_set = set()
pos_list = list()
panel_set = set()
for i, j, bg, bemi, pos, m in rows:
    print(i.value, j.value, bg.value, bemi.value, pos.value)
    bg_set.add(bg.value)
    bemi_set.add(bemi.value)
    temp_pos = []
#     if ws.cell(row=i.value, column =3) == bemi.value:
#         temp_pos.append(pos.value)
#     pos_list.insert(-1,temp_list)


print("BG Names: ", bg_set)
print('Bemi numbers: ', bemi_set)
print("position Number List: ", pos_list)



# Saving the file
#startmodel.save('C4-A20-{}_AFO{}_A00_{}.xlsx'.format(st_number[4:],afo,today))